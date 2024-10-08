import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as Oimage
from openpyxl.utils.units import pixels_to_points
from PIL import Image as PILImage, ImageDraw, ImageFont
import barcode
from barcode.writer import ImageWriter

# Constants
OUTPUT_EXCEL_PATH = 'data/tags_output.xlsx'
EXCEL_FILE_PATH = 'data/source_sheet.xlsx'
BARCODE_IMAGE_FOLDER = 'images/barcodes'
RUPEE_SYMBOL = "₹"
IMAGE_HEIGHT_INCHES = 0.84
IMAGE_WIDTH_INCHES = 1.83
DPI = 96  # DPI for image size calculation

# Create barcode images for each product
def generate_custom_barcode(barcode_number, BARCODE_IMAGE_FOLDER):
    # Create barcode filename based on last 4 digits
    file_name = f"{barcode_number[-4:]}.gif"
    final_image_path = os.path.join(BARCODE_IMAGE_FOLDER, file_name)
    
    # If barcode already exists, return the path
    if os.path.exists(final_image_path):
        return final_image_path
    
    # Barcode options for style and size
    options = {
        'module_width': 0.3,
        'module_height': 8,
        'text_distance': 0,
        'font_size': 0,
        'quiet_zone': 0.1,
        'write_text': False
    }
    
    # Generate the barcode and save as an image
    ean = barcode.get('Code 128', barcode_number, writer=ImageWriter())
    ean.write('barcode_image.png', options)
    
    # Load and add custom text to the barcode image
    barcode_image = PILImage.open('barcode_image.png')
    font = ImageFont.truetype("calibri.ttf", 25)
    
    # Create a new image canvas with room for human-readable text
    img_width, img_height = barcode_image.size
    new_image = PILImage.new('RGB', (img_width, img_height + 20), "white")
    new_image.paste(barcode_image, (0, 0))
    
    # Add text (barcode number) below the barcode
    draw = ImageDraw.Draw(new_image)
    text = barcode_number
    text_width, text_height = draw.textsize(text, font=font)
    text_x = (img_width - text_width) // 2
    text_y = img_height + 5
    draw.text((text_x, text_y), text, font=font, fill="black")
    
    # Save final barcode image
    new_image.save(final_image_path, format="GIF")
    
    return final_image_path

# Add a product tag to the Excel file
def add_tag_to_excel(sheet, price, description, mrp, size, barcode_image_path, row, col):
    sheet.column_dimensions[chr(65 + col)].width = 26.95

    # Define border style
    medium_border = Border(left=Side(style='medium'), right=Side(style='medium'),
                           top=Side(style='medium'), bottom=Side(style='medium'))

    # Top cell (price)
    cell = sheet.cell(row=row, column=col + 1)
    cell.value = f"{RUPEE_SYMBOL} {int(price)}"
    cell.font = Font(name='Calibri', size=40)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    sheet.row_dimensions[row].height = 42

    # Middle cell (barcode image)
    cell = sheet.cell(row=row + 1, column=col + 1)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[row + 1].height = 61.2

    if barcode_image_path:
        img = Oimage(barcode_image_path)
        img.width = int(IMAGE_WIDTH_INCHES * DPI)
        img.height = int(IMAGE_HEIGHT_INCHES * DPI)
        img.anchor = f"{chr(65 + col)}{row + 1}"
        sheet.add_image(img)

    # Bottom cell (description, MRP, size)
    description_text = f"Des: {description}\nMRP: {mrp}\nSize: {size}"
    cell = sheet.cell(row=row + 2, column=col + 1)
    cell.value = description_text
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.font = Font(size=11)
    sheet.row_dimensions[row + 2].height = 80.4

    # Apply borders to the whole tag
    for r in range(row, row + 3):
        for c in range(col + 1, col + 2):
            cell = sheet.cell(row=r, column=c)
            cell.border = medium_border

# Generate all barcodes from the Excel sheet
def generate_barcodes_from_excel(EXCEL_FILE_PATH, BARCODE_IMAGE_FOLDER):
    df = pd.read_excel(EXCEL_FILE_PATH)
    if 'Item Code' not in df.columns:
        raise ValueError("The 'Item Code' column does not exist in the Excel file.")
    
    for barcode_number in df['Item Code'].dropna():
        barcode_number_str = str(int(barcode_number))
        generate_custom_barcode(barcode_number_str, BARCODE_IMAGE_FOLDER)

# Main function to create tags in Excel
def main():
    # Ensure output directories exist
    os.makedirs(BARCODE_IMAGE_FOLDER, exist_ok=True)
    os.makedirs('data', exist_ok=True)

    # Load the Excel sheet
    df = pd.read_excel(EXCEL_FILE_PATH)

    # Create a new Excel workbook and set the sheet
    wb = Workbook()
    sheet = wb.active
    sheet.delete_rows(1)

    row, col = 1, 0
    tags_count = 0

    # Loop through rows to create tags
    for _, row_data in df.iterrows():
        barcode_number = str(int(row_data['Item Code']))
        barcode_image_name = barcode_number[-4:]
        barcode_image_path = os.path.join(BARCODE_IMAGE_FOLDER, f"{barcode_image_name}.gif")

        if not os.path.exists(barcode_image_path):
            generate_custom_barcode(barcode_number, BARCODE_IMAGE_FOLDER)

        # Add tags to Excel
        add_tag_to_excel(sheet, row_data['Selling Price'], row_data['Product Name'], 
                         row_data['MRP'], row_data['Size'], barcode_image_path, row, col)
        col += 1
        tags_count += 1

        # Move to the next row set after 7 tags
        if col == 7:
            col = 0
            row += 4

    # Save the Excel workbook
    output_excel_path = OUTPUT_EXCEL_PATH
    wb.save(output_excel_path)
    print(f"Tags saved to {output_excel_path}")

if __name__ == "__main__":
    main()
